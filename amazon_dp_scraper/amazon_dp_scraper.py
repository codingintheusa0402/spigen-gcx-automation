#!/usr/bin/env python3
"""
Amazon /dp/ Product Detail Scraper — async Playwright edition
- ONE Chrome window, N tabs (async Playwright, properly concurrent)
- Dual-language: Sheet 1 = English results | Sheet 2 = Local-language results
  (same headers in both sheets; non-English domains scraped twice)
- Scrapes: productTitle, brand, color, size, compatibility, rating, review_count,
  About this item, po-* attribute table, Item/Additional details tables
- Live xlsx write: rows appended to file as they complete
- Output: amazon_dp_YYYYMMDD_HHMMSS.xlsx  (two sheets inside)
- CLI: python3 amazon_dp_scraper.py [ASIN ...] [--domains US UK ...]
- Keyboard controls (macOS): ⌥P = pause   ⌥R = resume   ⌥Q = quit
"""

import asyncio, re, sys, argparse, threading, random
from datetime import datetime
from playwright.async_api import async_playwright, Browser, BrowserContext, Page

try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font
    _OPENPYXL = True
except ImportError:
    _OPENPYXL = False
    print('[warn] openpyxl not found — falling back to CSV output', flush=True)
    import csv

try:
    from pynput import keyboard as kb
    _PYNPUT = True
except ImportError:
    _PYNPUT = False

# ── Config ────────────────────────────────────────────────────────────────────

CONCURRENCY  = 6   # concurrent English tabs
LOCAL_TABS   = 2   # tabs per non-English language pool

DOMAINS = {
    'US': 'https://www.amazon.com',
    'UK': 'https://www.amazon.co.uk',
    'DE': 'https://www.amazon.de',
    'FR': 'https://www.amazon.fr',
    'IT': 'https://www.amazon.it',
    'ES': 'https://www.amazon.es',
    'JP': 'https://www.amazon.co.jp',
    'IN': 'https://www.amazon.in',
}

ENGLISH_DOMAINS = {'US', 'UK', 'IN'}

LOCAL_LANG = {
    'DE': 'de-DE,de;q=0.9',
    'FR': 'fr-FR,fr;q=0.9',
    'IT': 'it-IT,it;q=0.9',
    'ES': 'es-ES,es;q=0.9',
    'JP': 'ja-JP,ja;q=0.9',
}

_UA = (
    'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) '
    'AppleWebKit/537.36 (KHTML, like Gecko) '
    'Chrome/124.0.0.0 Safari/537.36'
)

PAGE_WAIT    = 2500   # ms after navigation
PAGE_TIMEOUT = 60000  # ms

# ── Keyboard controls (runs in a separate daemon thread) ──────────────────────

_paused   = threading.Event(); _paused.set()   # cleared = paused
_quit_flag = threading.Event()                  # set = stop after in-flight tasks

if _PYNPUT:
    _alt_down = False

    def _on_press(key):
        global _alt_down
        if key in (kb.Key.alt, kb.Key.alt_l, kb.Key.alt_r):
            _alt_down = True; return
        if not _alt_down: return
        char = getattr(key, 'char', None)
        vk   = getattr(key, 'vk',  None)
        if char in ('π', 'p', 'P') or vk == 35:
            _paused.clear()
            print('\n  ⏸  PAUSED — ⌥R to resume   ⌥Q to quit', flush=True)
        elif char in ('®', 'r', 'R') or vk == 15:
            _paused.set()
            print('\n  ▶  RESUMED', flush=True)
        elif char in ('œ', 'q', 'Q') or vk == 12:
            _quit_flag.set(); _paused.set()
            print('\n  ⏹  STOPPING...', flush=True)

    def _on_release(key):
        global _alt_down
        if key in (kb.Key.alt, kb.Key.alt_l, kb.Key.alt_r):
            _alt_down = False

    _kb = kb.Listener(on_press=_on_press, on_release=_on_release)
    _kb.daemon = True
    _kb.start()

async def _wait_if_paused():
    """Async-friendly pause: yields to event loop while paused."""
    while not _paused.is_set():
        await asyncio.sleep(0.3)

# ── Session pool ──────────────────────────────────────────────────────────────
# Each slot = isolated BrowserContext + Page.
# Separate contexts = separate sessions → Amazon treats each as a distinct user,
# preventing the rate-limiting that happens with a shared context.
# headless=True means zero visible windows regardless of context count.

class SessionPool:
    """Pool of (BrowserContext, Page) pairs — each slot is fully isolated."""
    def __init__(self, browser: Browser, size: int,
                 accept_language: str, locale: str = 'en-US'):
        self._browser = browser
        self._lang    = accept_language
        self._locale  = locale
        self._sem     = asyncio.Semaphore(size)
        self._q: asyncio.Queue = asyncio.Queue()

    async def _new_session(self):
        ctx = await self._browser.new_context(user_agent=_UA, locale=self._locale)
        await ctx.add_init_script(
            "Object.defineProperty(navigator,'webdriver',{get:()=>undefined})"
        )
        page = await ctx.new_page()
        await page.set_extra_http_headers({'Accept-Language': self._lang})
        return ctx, page

    async def init(self):
        for _ in range(self._sem._value):
            self._q.put_nowait(await self._new_session())

    async def acquire(self):
        await self._sem.acquire()
        return await self._q.get()   # returns (ctx, page)

    async def release(self, session):
        self._q.put_nowait(session)
        self._sem.release()

    async def replace(self, old_session):
        ctx, _ = old_session
        try: await ctx.close()
        except Exception: pass
        self._q.put_nowait(await self._new_session())
        self._sem.release()

    async def close_all(self):
        while not self._q.empty():
            try:
                ctx, _ = self._q.get_nowait()
                await ctx.close()
            except Exception: pass

# ── Popup dismissal ───────────────────────────────────────────────────────────

_POPUP_TEXTS = [
    'Accept', 'Accept All', 'Akzeptieren', 'Alle akzeptieren',
    'Accept Cookies', 'Accept all', '同意する',
]

async def dismiss_popups(page: Page):
    for text in _POPUP_TEXTS:
        try:
            btn = page.locator(
                f"button:text-matches('{text}', 'i'), "
                f"input[value*='{text}' i]"
            ).first
            if await btn.is_visible(timeout=500):
                await btn.click()
                await page.wait_for_timeout(400)
                return
        except Exception:
            pass

# ── Text helpers ──────────────────────────────────────────────────────────────

async def _t(page: Page, selector: str) -> str:
    try:
        return (await page.locator(selector).first.inner_text(timeout=2000)).strip()
    except Exception:
        return ''

async def _first(page: Page, selectors):
    for sel in selectors:
        v = await _t(page, sel)
        if v: return v
    return ''

def extract_rating(text):
    m = re.search(r'(\d[.,]\d)', text)
    return m.group(1).replace(',', '.') if m else ''

def extract_count(text):
    m = re.search(r'([\d][,\d\.]+\d|\d+)', text)
    return re.sub(r'[^\d]', '', m.group(1)) if m else ''

_PHONE_RE = re.compile(
    r'Galaxy\s+S\d+\s*(?:FE|Plus|\+|Ultra|Edge)?'
    r'|iPhone\s+\d+\s*(?:Pro\s*Max|Pro|Plus|Mini|e)?'
    r'|Pixel\s+\d+\s*(?:Pro\s*XL|Pro|a\s*XL|a|XL)?',
    re.I
)

def extract_size_from_title(t):
    m = _PHONE_RE.search(t); return m.group().strip() if m else ''

def extract_color_from_title(t):
    if ' - ' in t:
        c = t.rsplit(' - ', 1)[-1].strip()
        if c and len(c) < 40 and '[' not in c: return c
    return ''

# ── Page scrapers ─────────────────────────────────────────────────────────────

async def get_po_table(page: Page):
    result = {}
    try:
        for row in await page.locator(
            'table.a-normal.a-spacing-micro tr[class*="po-"]'
        ).all():
            try:
                cells = await row.locator('td').all()
                if len(cells) >= 2:
                    k = (await cells[0].inner_text(timeout=1000)).strip()
                    v = (await cells[1].inner_text(timeout=1000)).strip()
                    if k and v: result[k] = v
            except Exception: pass
    except Exception: pass
    return result

async def get_about_item(page: Page):
    items = []
    try:
        for el in await page.locator('#feature-bullets li span.a-list-item').all():
            txt = (await el.inner_text(timeout=1000)).strip()
            if txt: items.append(txt)
    except Exception: pass
    return ' | '.join(items)

_PRODDET_SKIP = frozenset({
    'customer reviews', 'best sellers rank', 'unspsc code',
    'customer reviews (including product star ratings)',
})

async def expand_and_get_proddet(page: Page):
    for heading in ('Item details', 'Additional details'):
        try:
            hdr = page.locator(
                f"a.a-expander-header:has(span:text-is('{heading}'))"
            ).first
            if await hdr.get_attribute('aria-expanded', timeout=1000) != 'true':
                await hdr.evaluate('el => el.click()')
                await page.wait_for_timeout(500)
        except Exception: pass
    result = {}
    try:
        for table in await page.locator('table.a-keyvalue.prodDetTable').all():
            for row in await table.locator('tr').all():
                try:
                    th = (await row.locator('th').first.inner_text(timeout=1000)).strip()
                    td = (await row.locator('td').first.inner_text(timeout=1000)).strip()
                    if th and td and th.lower() not in _PRODDET_SKIP:
                        result[th] = td
                except Exception: pass
    except Exception: pass
    return result

async def get_detail_table(page: Page):
    result = {}
    try:
        for row in await page.locator(
            '#productDetails_techSpec_section_1 tr, '
            '#productDetails_detailBullets_sections1 tr, '
            '#detailBullets_feature_div li'
        ).all():
            try:
                text = await row.inner_text(timeout=1000)
                if ':' in text:
                    parts = text.split(':', 1)
                    k = parts[0].strip().lower(); v = parts[1].strip()
                    if k and v: result[k] = v
            except Exception: pass
    except Exception: pass
    for sel in ('tr.po-brand', 'tr.po-color', 'tr.po-size'):
        try:
            for row in await page.locator(sel).all():
                cells = await row.locator('td').all()
                if len(cells) >= 2:
                    k = (await cells[0].inner_text(timeout=500)).strip().lower()
                    v = (await cells[1].inner_text(timeout=500)).strip()
                    if k and v: result[k] = v
        except Exception: pass
    return result

def _find(table, *keys):
    for k in keys:
        for tk in table:
            if k.lower() in tk: return table[tk]
    return ''

async def get_compatibility(page: Page):
    try:
        for b in await page.locator('#feature-bullets li span.a-list-item').all():
            txt = await b.inner_text(timeout=1000)
            if re.search(r'compat|fit[s ]|work[s ]|design|for (galaxy|iphone|pixel|samsung)', txt, re.I):
                return txt.strip()
    except Exception: pass
    return ''

# ── Core scrape ───────────────────────────────────────────────────────────────

async def _is_robot_check(page: Page) -> bool:
    """Detect Amazon's 'continue shopping' / 'enter the characters' bot-check page."""
    try:
        body = await page.locator('body').inner_text(timeout=2000)
        b = body.lower()
        return ('continue shopping' in b or
                'enter the characters' in b or
                'type the characters' in b or
                'robot or human' in b)
    except Exception:
        return False


async def scrape_page(page: Page, asin: str, domain_key: str, domain_url: str) -> dict:
    url = f"{domain_url}/dp/{asin}"
    await asyncio.sleep(random.uniform(0.3, 1.5))
    await page.goto(url, wait_until='domcontentloaded', timeout=PAGE_TIMEOUT)
    await page.wait_for_timeout(PAGE_WAIT)

    # If Amazon shows a bot-check page, warm up the session via the homepage first
    if await _is_robot_check(page):
        await page.goto(domain_url, wait_until='domcontentloaded', timeout=PAGE_TIMEOUT)
        await page.wait_for_timeout(2000)
        await dismiss_popups(page)
        await page.goto(url, wait_until='domcontentloaded', timeout=PAGE_TIMEOUT)
        await page.wait_for_timeout(PAGE_WAIT)

    await dismiss_popups(page)

    product_title = await _t(page, '#productTitle') or await _t(page, 'h1#title span')

    brand = await _first(page, [
        '#bylineInfo', 'a#bylineInfo',
        '.po-brand .po-break-word', 'tr.po-brand td:last-child span',
    ])
    if brand.lower().startswith('visit the '):  brand = brand[10:].strip()
    if brand.lower().startswith('brand: '):     brand = brand[7:].strip()

    rating = extract_rating(await _first(page, [
        'span[data-hook="rating-out-of-text"]',
        '#averageCustomerReviews span.a-icon-alt',
        'i[data-hook="average-star-rating"] span.a-icon-alt',
    ]))
    review_count = extract_count(await _first(page, [
        '#acrCustomerReviewText',
        'span[data-hook="total-review-count"]',
        '#acrCustomerReviewLink span',
    ]))

    color = await _first(page, [
        '#variation_color_name option.dropdownSelect',
        '#variation_color_name span.a-dropdown-prompt',
        '#variation_color_name .selection',
        '#variation_color_name span.a-color-base',
    ])
    size = await _first(page, [
        '#variation_size_name option.dropdownSelect',
        '#variation_size_name span.a-dropdown-prompt',
        '#variation_size_name .selection',
        '#variation_configuration_name option.dropdownSelect',
        '#variation_configuration_name span.a-dropdown-prompt',
        '#variation_configuration_name .selection',
    ])

    po_table     = await get_po_table(page)
    about_item   = await get_about_item(page)
    proddet      = await expand_and_get_proddet(page)
    detail_table = await get_detail_table(page)

    if not color:   color   = _find(detail_table, 'color')
    if not size:    size    = _find(detail_table, 'size', 'configuration')
    if not brand:   brand   = _find(detail_table, 'brand', 'manufacturer')
    if not size:    size    = extract_size_from_title(product_title)
    if not color:   color   = extract_color_from_title(product_title)

    compatibility = await get_compatibility(page)
    if not compatibility: compatibility = _find(detail_table, 'compat', 'fit')

    row = {
        'domain':          domain_key,
        'asin':            asin,
        'url':             url,
        'productTitle':    product_title,
        'brand':           brand,
        'color':           color,
        'size':            size,
        'compatibility':   compatibility,
        'About this item': about_item,
        'rating':          rating,
        'review_count':    review_count,
        'scraped_at':      datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
    }
    row.update(po_table)
    row.update(proddet)
    return row

# ── xlsx live writer ──────────────────────────────────────────────────────────

BASE_FIELDS = [
    'domain', 'asin', 'url', 'productTitle', 'brand', 'color', 'size',
    'compatibility', 'About this item', 'rating', 'review_count', 'scraped_at',
]

class XlsxWriter:
    """
    Keeps workbook in memory. On each new row, if no new columns appear it
    appends efficiently; if new columns are found it rewrites both sheets.
    Thread-safe (asyncio lock).
    """
    def __init__(self, path: str, sheet_en: str, sheet_locale: str):
        self._path = path
        self._lock = asyncio.Lock()
        self._headers = list(BASE_FIELDS)   # grows as new columns appear
        self._rows_en     = []
        self._rows_locale = []

        if _OPENPYXL:
            self._wb = openpyxl.Workbook()
            self._ws_en = self._wb.active
            self._ws_en.title = sheet_en
            self._ws_locale = self._wb.create_sheet(sheet_locale)
            self._write_header_row(self._ws_en)
            self._write_header_row(self._ws_locale)
            self._wb.save(path)
        else:
            self._csv_en     = open(path.replace('.xlsx','_en.csv'),     'w', newline='', encoding='utf-8')
            self._csv_locale = open(path.replace('.xlsx','_locale.csv'), 'w', newline='', encoding='utf-8')
            self._wr_en     = None
            self._wr_locale = None

    def _write_header_row(self, ws):
        for c, h in enumerate(self._headers, 1):
            cell = ws.cell(row=1, column=c, value=h)
            cell.font = Font(bold=True)
            cell.fill = PatternFill('solid', fgColor='D9E1F2')

    def _rewrite_sheet(self, ws, rows):
        ws.delete_rows(1, ws.max_row)
        self._write_header_row(ws)
        for r_idx, row in enumerate(rows, 2):
            for c_idx, h in enumerate(self._headers, 1):
                ws.cell(row=r_idx, column=c_idx, value=row.get(h, ''))

    def _append_sheet_row(self, ws, row):
        r_idx = ws.max_row + 1
        for c_idx, h in enumerate(self._headers, 1):
            ws.cell(row=r_idx, column=c_idx, value=row.get(h, ''))

    async def add(self, row: dict, is_locale: bool):
        async with self._lock:
            new_cols = [k for k in row if k not in self._headers]
            if new_cols:
                self._headers.extend(sorted(new_cols))
                if _OPENPYXL:
                    # Rewrite both sheets with expanded headers
                    self._rewrite_sheet(self._ws_en, self._rows_en)
                    self._rewrite_sheet(self._ws_locale, self._rows_locale)

            if is_locale:
                self._rows_locale.append(row)
                if _OPENPYXL and not new_cols:
                    self._append_sheet_row(self._ws_locale, row)
            else:
                self._rows_en.append(row)
                if _OPENPYXL and not new_cols:
                    self._append_sheet_row(self._ws_en, row)

            if _OPENPYXL:
                self._wb.save(self._path)
            else:
                # CSV fallback
                all_rows = (self._rows_en, is_locale) if not is_locale else (self._rows_locale, is_locale)
                # simple rewrite
                pass  # handled below

    def final_save(self):
        if _OPENPYXL:
            # Final rewrite to ensure ordering is correct
            self._rewrite_sheet(self._ws_en, self._rows_en)
            self._rewrite_sheet(self._ws_locale, self._rows_locale)
            self._wb.save(self._path)
        else:
            def _write(path, rows):
                with open(path, 'w', newline='', encoding='utf-8') as f:
                    w = csv.DictWriter(f, fieldnames=self._headers, extrasaction='ignore', restval='')
                    w.writeheader(); w.writerows(rows)
            _write(self._path.replace('.xlsx','_en.csv'),     self._rows_en)
            _write(self._path.replace('.xlsx','_locale.csv'), self._rows_locale)

    def stats(self):
        return len(self._rows_en), len(self._rows_locale), len(self._headers)

# ── Worker ────────────────────────────────────────────────────────────────────

def _empty(domain_key, asin, domain_url):
    return {
        'domain': domain_key, 'asin': asin, 'url': f"{domain_url}/dp/{asin}",
        'productTitle': '', 'brand': '', 'color': '', 'size': '',
        'compatibility': '', 'About this item': '',
        'rating': '', 'review_count': '',
        'scraped_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
    }

async def worker(
    task, en_pool: SessionPool, local_pools: dict,
    writer: XlsxWriter, counter: list, total: int
):
    domain_key, domain_url, asin = task
    if _quit_flag.is_set(): return
    await _wait_if_paused()
    if _quit_flag.is_set(): return

    # ── English scrape ────────────────────────────────────────────────────────
    session = await en_pool.acquire()
    try:
        _, page = session
        row_en = await scrape_page(page, asin, domain_key, domain_url)
        await en_pool.release(session)
    except Exception:
        await en_pool.replace(session)
        session2 = await en_pool.acquire()
        try:
            _, page2 = session2
            row_en = await scrape_page(page2, asin, domain_key, domain_url)
            await en_pool.release(session2)
        except Exception:
            await en_pool.replace(session2)
            row_en = _empty(domain_key, asin, domain_url)

    await writer.add(row_en, is_locale=False)

    # ── Local-language scrape (non-English domains only) ──────────────────────
    # Always attempt regardless of English result — the two scrapes are independent.
    # English may be blocked while local succeeds, or vice versa.
    if domain_key not in ENGLISH_DOMAINS and domain_key in local_pools:
        lpool = local_pools[domain_key]
        lsession = await lpool.acquire()
        try:
            _, lpage = lsession
            row_loc = await scrape_page(lpage, asin, domain_key, domain_url)
            await lpool.release(lsession)
            await writer.add(row_loc, is_locale=True)
        except Exception:
            await lpool.replace(lsession)

    counter[0] += 1
    done = counter[0]
    print(f"  [{done}/{total}] {domain_key} {asin}  "
          f"rating={row_en.get('rating','') or '-'}  "
          f"reviews={row_en.get('review_count','') or '-'}  "
          f"size={str(row_en.get('size',''))[:20] or '-'}", flush=True)

# ── Main ──────────────────────────────────────────────────────────────────────

def parse_args():
    p = argparse.ArgumentParser(description='Amazon /dp/ scraper')
    p.add_argument('asins', nargs='*', help='ASINs to scrape (default: built-in list)')
    p.add_argument('--domains', nargs='+', metavar='CC',
                   help='Domain codes, e.g. --domains US UK DE (default: all 8)')
    return p.parse_args()

RAW_ASINS = """
B0GDHXBK25 B0FVB2PF8R B0FXCJG9L9 B0FVBMZYJX B0FV9TK5KH B0FVB24LQ4
B0FVBKXV4Z B0FVBMG3B5 B0FVB2D675 B0FVBKZCBC B0FV9YH7B3 B0FWGVKRW9
B0FWGDN69P B0FVB8NGJN B0FVB9TM8Q B0FVBMP4Y6 B0FVBKSRYL B0FWGJBGQS
B0FVBFXHSW B0FVBMQ4MC B0FVB8QL42 B0FV9ZHSN7 B0FVBN27XW B0FVB4W8WR
KRACS10727 KRACS11057 KRACS11058
B0GDHRZFWH B0FV9WPXMS B0FXCHLSRG B0FVBHQNFB B0FVBMCJBQ B0FVB2TNPR
B0FV9RK8P8 B0FVB7135D B0FVB3W7W3 B0FVB1ZVBC B0FVBBZ1RJ B0FV9ZGNMN
B0FWGL3F69 B0FWG5C6SQ B0FVB661VQ B0FVB23HNN B0FVB1GM81 B0FVB1BBZD
B0FWH9N11J B0FVB63K45 B0FVB67FK4 B0FV9TH2FC
KRACS10701 KRACS11046 KRACS11047
B0GDHBDG47 B0FVB1GXJ7 B0FXCK6S9B B0FVBMCTG6 B0FVBF3CDJ B0FVB84WZK
B0FVBFT41L B0FVB24LQ3 B0FVBMVTRH B0FVB57R9R B0FVBKG9ZP B0FVBMNSG3
B0FV9ZHFMT B0FVBKD8N9 B0FWGPKRH5 B0FWG9RXZF B0FVBPTQT2 B0FVBHT64R
B0FVB5V1MQ B0FVB5CDWL B0FWGWJRD3 B0FVB4P3R1 B0FXCK8P3P B0FV9N1Y5H
B0FVB569XJ B0FVBFF7BJ B0FVBPV3QF B0FVBPVR1M B0FXCLF5L3 B0FVBKHYBR
B0FVB2RXM3 B0FVB22GHK
KRACS10671 KRACS11027 KRACS11028
B0G7QZ7RMD B0G7RYP439 B0G7RJH146 B0G6CFQBJJ B0G6CQ5JPH B0G7RD3G98
KRAGL11080
B0G7Q4KVJJ B0G7RK63K6 B0G6CLBFDB B0G6CLKRFR B0G7S8J7SJ B0G877D436
B0G7STMFKY B0G6CG1B7L
""".split()

DEFAULT_ASINS = [a for a in RAW_ASINS if not a.startswith('KR')]


async def main():
    args = parse_args()
    asins  = args.asins if args.asins else DEFAULT_ASINS
    d_keys = set(args.domains) if args.domains else set(DOMAINS.keys())
    active = {k: v for k, v in DOMAINS.items() if k in d_keys}

    if not asins:  print('No ASINs.', flush=True); return
    if not active: print('No domains.', flush=True); return

    tasks = [(dk, du, asin) for dk, du in active.items() for asin in asins]
    total = len(tasks)

    ts   = datetime.now().strftime('%Y%m%d_%H%M%S')
    path = f"/Users/kevinkim/Desktop/amazon_dp_{ts}.xlsx"
    sheet_en     = f"amazon_dp_{ts}"          # 25 chars — within Excel's 31-char limit
    sheet_locale = f"amazon_dp_{ts}_loc"      # 29 chars — within limit

    foreign_keys = [dk for dk in active if dk in LOCAL_LANG]
    n_en_tabs    = CONCURRENCY
    n_local_tabs = LOCAL_TABS * len(foreign_keys)

    print(f"\n  {len(asins)} ASINs × {len(active)} domains = {total} pages", flush=True)
    print(f"  Concurrency: {n_en_tabs} English tabs + {n_local_tabs} local-language tabs", flush=True)
    print(f"  All tabs in ONE Chrome window", flush=True)
    print(f"  Output: {path}", flush=True)
    if _PYNPUT:
        print(f"  Controls: ⌥P pause   ⌥R resume   ⌥Q quit\n", flush=True)

    writer  = XlsxWriter(path, sheet_en, sheet_locale)
    counter = [0]

    async with async_playwright() as pw:
        # headless=True: no visible windows.
        # Each session pool slot gets its own BrowserContext → isolated Amazon sessions.
        browser = await pw.chromium.launch(
            headless=True,
            args=[
                '--no-sandbox',
                '--disable-dev-shm-usage',
                '--disable-blink-features=AutomationControlled',
            ],
        )

        print(f"  Creating {n_en_tabs} English sessions...", flush=True)
        en_pool = SessionPool(browser, n_en_tabs, 'en-US,en;q=0.9', locale='en-US')
        await en_pool.init()

        local_pools = {}
        for dk in foreign_keys:
            lang   = LOCAL_LANG[dk]
            locale = lang.split(',')[0].split(';')[0]   # e.g. 'de-DE'
            print(f"  Creating {LOCAL_TABS} {dk} sessions ({lang})...", flush=True)
            lp = SessionPool(browser, LOCAL_TABS, lang, locale=locale)
            await lp.init()
            local_pools[dk] = lp

        print(f"  All {n_en_tabs + n_local_tabs} sessions ready — starting scrape\n", flush=True)

        sem = asyncio.Semaphore(CONCURRENCY)

        async def bounded(task):
            async with sem:
                await worker(task, en_pool, local_pools, writer, counter, total)

        try:
            await asyncio.gather(*[bounded(t) for t in tasks])
        except asyncio.CancelledError:
            pass
        finally:
            await en_pool.close_all()
            for lp in local_pools.values():
                await lp.close_all()
            await browser.close()

    writer.final_save()
    n_en, n_loc, n_cols = writer.stats()
    print(f"\n  Done — {n_en} English rows + {n_loc} locale rows × {n_cols} columns → {path}", flush=True)


if __name__ == '__main__':
    asyncio.run(main())
